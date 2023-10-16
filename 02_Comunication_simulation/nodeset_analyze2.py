from collections import defaultdict
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook, load_workbook

NODE_SET_FILE_PATH = 'CNC\Opc.Ua.CNC.NodeSet.xml'
RESULT_FILE_PATH = 'CNC\Analyze.xlsx'

# Parse the XML file
tree = ET.parse(NODE_SET_FILE_PATH)
root = tree.getroot()

# Extract NodeId number
def extract_nodeid(nodeid):
    return nodeid.split('=')[-1]

# Extract namespace and reference ID
def extract_namespace_reference(reference):
    parts = reference.split(';')
    if len(parts) == 1:  # No namespace specified, default to "1" #! default to "0"
        namespace = '0' 
        reference_id = parts[0].split('=')[-1]
    else:  # Namespace specified
        namespace = parts[0].split('=')[-1]
        reference_id = parts[1].split('=')[-1]
    return namespace, reference_id

# Extract references
def extract_references(references_element):
    references = {}
    for reference in references_element:
        namespace, reference_id = extract_namespace_reference(reference.text)
        if namespace in references:
            references[namespace].append(reference_id)
        else:
            references[namespace] = [reference_id]
    return references

# Prepare for CSV writing
csv_data = []
header = ["UA Type", "DisplayName", "NodeId", "References", "Description"]


# Iterate over all elements in the XML tree
for element in root.iter():
    if 'UAObjectType' in element.tag or 'UAVariable' in element.tag or 'UAObject' in element.tag:
        ua_type = element.tag.split('}')[-1]
        display_name = element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}DisplayName').text
        nodeid = extract_nodeid(element.attrib.get('NodeId', ''))
        description_element = element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}Description')
        description = description_element.text if description_element is not None else ''
        references_element = element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}References')
        references = extract_references(references_element) if references_element is not None else {}
        # Convert references dict to string
        references_str = '; '.join([f"{ns}:[{', '.join(ids)}]" for ns, ids in references.items()])
        csv_data.append([ua_type, display_name, nodeid, references_str, description])

# Step 1: Build a tree structure
tree = defaultdict(list)
node_names = {}

# Iterate over all elements in the XML tree
for element in root.iter():
    if 'UAObjectType' in element.tag or 'UAVariable' in element.tag or 'UAObject' in element.tag:
        nodeid = extract_nodeid(element.attrib.get('NodeId', ''))
        parent_nodeid = extract_nodeid(element.attrib.get('ParentNodeId', '')) if 'ParentNodeId' in element.attrib else None
        display_name = element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}DisplayName').text
        node_names[nodeid] = display_name  # Store node names for later use
        if parent_nodeid:
            tree[parent_nodeid].append(nodeid)  # Add child to parent in tree

# Step 2: Perform a depth-first traversal and record the paths
paths = []

def dfs(nodeid, path):
    # Record the path
    paths.append(path)
    # Visit all children
    for child_nodeid in tree[nodeid]:
        dfs(child_nodeid, path + [node_names[child_nodeid]])

# Start traversal from each root (a node without a parent)
for nodeid in node_names:
    # if nodeid not in tree:  # No children, skip
    #     continue
    if any(nodeid in children for children in tree.values()):  # This node is a child of another node, skip
        continue
    dfs(nodeid, [node_names[nodeid]])
        
# Prepare DataFrame for Node Relation Table
node_table_df = pd.DataFrame(csv_data, columns=header)
node_relation_df = pd.DataFrame(paths)

# Create an Excel writer object
excel_writer = pd.ExcelWriter(RESULT_FILE_PATH, engine='openpyxl')

# Write DataFrames to Excel with different sheets
node_table_df.to_excel(excel_writer, sheet_name='Node Table', index=False)
node_relation_df.to_excel(excel_writer, sheet_name='Node Relation', index=False)

# Save the Excel file
excel_writer.save()

# Open the Excel file using openpyxl
workbook = load_workbook(RESULT_FILE_PATH)

# Function to merge adjacent vertical cells with the same content
def merge_cells(sheet):
    for col in sheet.iter_cols():
        start_idx = None
        for idx, cell in enumerate(col):
            # Check if the current cell has the same value as the previous cell
            if start_idx is not None and (cell.value != col[start_idx].value or cell.value is None):
                if idx - start_idx > 1: # Merge only if there are more than one adjacent cells with the same value
                    sheet.merge_cells(start_row=start_idx + 1, start_column=cell.column, 
                                      end_row=idx, end_column=cell.column)
                start_idx = None
            # If the current cell has a value and is the same as the previous cell, mark the start index
            if cell.value is not None and (start_idx is None or cell.value == col[start_idx].value):
                if start_idx is None:
                    start_idx = idx

# Merge cells
merge_cells(workbook['Node Table'])
merge_cells(workbook['Node Relation'])

# Save the changes to the Excel file
workbook.save(RESULT_FILE_PATH)