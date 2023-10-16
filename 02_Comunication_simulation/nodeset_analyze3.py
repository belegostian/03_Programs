from collections import defaultdict
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook, load_workbook

NODE_SET_FILE_PATH = 'CNC\Opc.Ua.CNC.NodeSet.xml'
RESULT_FILE_PATH = 'CNC\Analyze.xlsx'

#* Parsing the XML File

def extract_nodeid(nodeid):
    return nodeid.split('=')[-1]

def extract_namespace_reference(reference):
    parts = reference.split(';')
    if len(parts) == 1:  # No namespace specified, default to "0"
        namespace = '0'
        reference_id = parts[0].split('=')[-1]
    else:  # Namespace specified
        namespace = parts[0].split('=')[-1]
        reference_id = parts[1].split('=')[-1]
    return namespace, reference_id

def extract_references(references_element):
    references = {}
    for reference in references_element:
        namespace, reference_id = extract_namespace_reference(reference.text)
        if namespace in references:
            references[namespace].append(reference_id)
        else:
            references[namespace] = [reference_id]
    return references

def extract_description(element):
    description_element = element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}Description')
    return description_element.text if description_element is not None else ''

def extract_element_data(element):
    ua_type = element.tag.split('}')[-1]
    display_name = element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}DisplayName').text
    nodeid = extract_nodeid(element.attrib.get('NodeId', ''))
    description = extract_description(element)
    references = extract_references(element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}References'))
    references_str = '; '.join([f"{ns}:[{', '.join(ids)}]" for ns, ids in references.items()])
    return [ua_type, display_name, nodeid, references_str, description]

def parse_xml():
    tree = ET.parse(NODE_SET_FILE_PATH)
    root = tree.getroot()
    csv_data = []
    header = ["UA Type", "DisplayName", "NodeId", "References", "Description"]

    for element in root.iter():
        if any(tag in element.tag for tag in ['UAObjectType', 'UAVariable', 'UAObject']):
            csv_data.append(extract_element_data(element))

    return csv_data, header, root

#* Building the Tree Structure

def extract_tree_data(element):
    nodeid = extract_nodeid(element.attrib.get('NodeId', ''))
    parent_nodeid = extract_nodeid(element.attrib.get('ParentNodeId', '')) if 'ParentNodeId' in element.attrib else None
    display_name = element.find('{http://opcfoundation.org/UA/2011/03/UANodeSet.xsd}DisplayName').text
    return nodeid, parent_nodeid, display_name

def build_tree(root):
    tree = defaultdict(list)
    node_names = {}

    for element in root.iter():
        if any(tag in element.tag for tag in ['UAObjectType', 'UAVariable', 'UAObject']):
            nodeid, parent_nodeid, display_name = extract_tree_data(element)
            node_names[nodeid] = display_name
            if parent_nodeid:
                tree[parent_nodeid].append(nodeid)

    return tree, node_names

def dfs(nodeid, path, tree, node_names, paths):
    paths.append(path)
    for child_nodeid in tree[nodeid]:
        dfs(child_nodeid, path + [node_names[child_nodeid]], tree, node_names, paths)


def traverse_tree(tree, node_names):
    paths = []
    for nodeid in node_names:
        if not any(nodeid in children for children in tree.values()):
            dfs(nodeid, [node_names[nodeid]], tree, node_names, paths)
    return paths

def write_to_excel(csv_data, header, paths):
    node_table_df = pd.DataFrame(csv_data, columns=header)
    node_relation_df = pd.DataFrame(paths)

    with pd.ExcelWriter(RESULT_FILE_PATH, engine='openpyxl') as excel_writer:
        node_table_df.to_excel(excel_writer, sheet_name='Node Table', index=False)
        node_relation_df.to_excel(excel_writer, sheet_name='Node Relation', index=False)

    workbook = load_workbook(RESULT_FILE_PATH)
    merge_cells(workbook['Node Table'])
    merge_cells(workbook['Node Relation'])
    workbook.save(RESULT_FILE_PATH)


def merge_cells(sheet):
    for col in sheet.iter_cols():
        start_idx = None
        for idx, cell in enumerate(col):
            if start_idx is not None and (cell.value != col[start_idx].value or cell.value is None):
                if idx - start_idx > 1:
                    sheet.merge_cells(start_row=start_idx + 1, start_column=cell.column,
                                      end_row=idx, end_column=cell.column)
                start_idx = None
            if cell.value is not None and (start_idx is None or cell.value == col[start_idx].value):
                if start_idx is None:
                    start_idx = idx


if __name__ == "__main__":
    csv_data, header, root = parse_xml()
    tree_structure, node_names = build_tree(root)
    paths = traverse_tree(tree_structure, node_names)
    write_to_excel(csv_data, header, paths)